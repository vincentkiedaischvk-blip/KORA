"""Erzeugt business/KORA-Abomodell.xlsx.

Die Arbeitsmappe enthält Formeln, keine berechneten Werte — wer eine Annahme
ändert, sieht die Projektion neu. Dieselbe Rechnung liegt als reines Python in
scripts/build-data.py und landet in data/unit-economics.csv, damit sie ohne
Excel prüfbar ist.
"""

import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "business", "KORA-Abomodell.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
H1 = Font(name="Arial", size=16, bold=True)
H2 = Font(name="Arial", size=11, bold=True, color="FFFFFF")
YELLOW = PatternFill("solid", fgColor="FFFF00")
DARK = PatternFill("solid", fgColor="0D1219")
GREYF = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="D9D9D9")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

EUR = '€#,##0;(€#,##0);-'
EUR2 = '€#,##0.00'
PCT = '0.0%'
NUM = '#,##0'

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════
# BLATT 1 — ANNAHMEN
# ═══════════════════════════════════════════════════════════
a = wb.active
a.title = "Annahmen"
a["A1"] = "KORA Safety — Abo-Modell Anbieterseite"
a["A1"].font = H1
a["A2"] = "Alle blauen Zellen sind Eingaben. Schwarze Zellen sind Formeln — nicht überschreiben."
a["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")

rows = [
    ("MARKT", None, None, None),
    ("Wach- und Sicherheitsunternehmen in DE", 5330, NUM,
     "Destatis, WZ 80.1, Stand Sept. 2023"),
    ("davon Top 25 (eigener Vertrieb, nicht adressierbar)", 25, NUM,
     "Lünendonk-Liste 2026"),
    ("Realistisch adressierbar (Anteil)", 0.45, PCT,
     "ANNAHME: Betriebe mit Event-/Kurzfristgeschäft. Rahmenvertrags-Werkschutz kommt nicht auf eine Plattform."),
    ("Zielgruppe gesamt", None, NUM, "Formel"),
    ("", None, None, None),
    ("PREIS", None, None, None),
    ("Abopreis pro Anbieter / Monat", 99, EUR2, "Netto, zzgl. USt."),
    ("Jahresumsatz pro Anbieter", None, EUR, "Formel"),
    ("", None, None, None),
    ("WACHSTUM", None, None, None),
    ("Neukunden Monat 1", 5, NUM, "ANNAHME: manuelle Akquise per Telefon"),
    ("Monatliches Wachstum der Neukunden", 0.15, PCT,
     "ANNAHME: 15 % mehr Neuabschlüsse pro Monat"),
    ("Obergrenze Neukunden / Monat", 60, NUM,
     "ANNAHME: Kapazitätsgrenze ohne größeres Vertriebsteam"),
    ("", None, None, None),
    ("ABWANDERUNG", None, None, None),
    ("Monatliche Churn-Rate", 0.04, PCT,
     "ANNAHME: 4 % im KMU-Segment bei monatlicher Kündbarkeit. Kritischste Zahl im Modell."),
    ("Rechnerische Kundenlebensdauer (Monate)", None, "0.0", "Formel: 1 / Churn"),
    ("Lifetime Value je Anbieter", None, EUR, "Formel"),
    ("", None, None, None),
    ("AKQUISE", None, None, None),
    ("Kosten pro Neukunde (CAC)", 250, EUR,
     "ANNAHME: Mix aus Eigenakquise und Paid Ads"),
    ("Amortisationsdauer (Monate)", None, "0.0", "Formel: CAC / Preis"),
    ("LTV / CAC", None, "0.0x", "Formel. Unter 3,0 ist das Modell nicht tragfähig."),
]

r = 4
labels = {}
for label, val, fmt, note in rows:
    if label and val is None and fmt is None:
        c = a.cell(row=r, column=1, value=label)
        c.font = H2
        c.fill = DARK
        for cc in range(1, 5):
            a.cell(row=r, column=cc).fill = DARK
    elif label:
        a.cell(row=r, column=1, value=label).font = BLACK
        cell = a.cell(row=r, column=2)
        if val is not None:
            cell.value = val
            cell.font = BLUE
            cell.fill = YELLOW
        else:
            cell.font = BLACK
        if fmt:
            cell.number_format = fmt
        cell.border = BORD
        a.cell(row=r, column=3, value=note).font = Font(name="Arial", size=9, color="666666")
        labels[label] = r
    r += 1

# Formeln
a[f"B{labels['Zielgruppe sesamt'] if False else labels['Zielgruppe gesamt']}"] = \
    f"=ROUND((B{labels['Wach- und Sicherheitsunternehmen in DE']}-B{labels['davon Top 25 (eigener Vertrieb, nicht adressierbar)']})*B{labels['Realistisch adressierbar (Anteil)']},0)"
a[f"B{labels['Jahresumsatz pro Anbieter']}"] = f"=B{labels['Abopreis pro Anbieter / Monat']}*12"
a[f"B{labels['Rechnerische Kundenlebensdauer (Monate)']}"] = f"=1/B{labels['Monatliche Churn-Rate']}"
a[f"B{labels['Lifetime Value je Anbieter']}"] = \
    f"=B{labels['Abopreis pro Anbieter / Monat']}*B{labels['Rechnerische Kundenlebensdauer (Monate)']}"
a[f"B{labels['Amortisationsdauer (Monate)']}"] = \
    f"=B{labels['Kosten pro Neukunde (CAC)']}/B{labels['Abopreis pro Anbieter / Monat']}"
a[f"B{labels['LTV / CAC']}"] = \
    f"=B{labels['Lifetime Value je Anbieter']}/B{labels['Kosten pro Neukunde (CAC)']}"

a.column_dimensions["A"].width = 46
a.column_dimensions["B"].width = 16
a.column_dimensions["C"].width = 78

P = f"Annahmen!$B${labels['Abopreis pro Anbieter / Monat']}"
CHURN = f"Annahmen!$B${labels['Monatliche Churn-Rate']}"
NEW1 = f"Annahmen!$B${labels['Neukunden Monat 1']}"
GROW = f"Annahmen!$B${labels['Monatliches Wachstum der Neukunden']}"
CAP = f"Annahmen!$B${labels['Obergrenze Neukunden / Monat']}"
CACC = f"Annahmen!$B${labels['Kosten pro Neukunde (CAC)']}"
TARGET = f"Annahmen!$B${labels['Zielgruppe gesamt']}"

# ═══════════════════════════════════════════════════════════
# BLATT 2 — PROJEKTION 36 MONATE
# ═══════════════════════════════════════════════════════════
p = wb.create_sheet("Projektion")
p["A1"] = "36-Monats-Projektion"
p["A1"].font = H1

HEAD = ["Monat", "Neukunden", "Kündigungen", "Aktive Abos", "MRR", "ARR",
        "Akquisekosten", "Deckungsbeitrag kumuliert", "Marktdurchdringung"]
for c, h in enumerate(HEAD, 1):
    cell = p.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    cell.fill = DARK
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
p.row_dimensions[3].height = 34

for m in range(1, 37):
    r = 3 + m
    p.cell(row=r, column=1, value=m).font = BOLD
    # Neukunden mit Deckel
    p.cell(row=r, column=2, value=f"=MIN({CAP},ROUND({NEW1}*(1+{GROW})^({m}-1),0))")
    # Kündigungen auf Bestand des Vormonats
    if m == 1:
        p.cell(row=r, column=3, value=0)
        p.cell(row=r, column=4, value=f"=B{r}-C{r}")
    else:
        p.cell(row=r, column=3, value=f"=ROUND(D{r-1}*{CHURN},0)")
        p.cell(row=r, column=4, value=f"=D{r-1}+B{r}-C{r}")
    p.cell(row=r, column=5, value=f"=D{r}*{P}")
    p.cell(row=r, column=6, value=f"=E{r}*12")
    p.cell(row=r, column=7, value=f"=B{r}*{CACC}")
    if m == 1:
        p.cell(row=r, column=8, value=f"=E{r}-G{r}")
    else:
        p.cell(row=r, column=8, value=f"=H{r-1}+E{r}-G{r}")
    p.cell(row=r, column=9, value=f"=D{r}/{TARGET}")

    for c in range(1, 10):
        cell = p.cell(row=r, column=c)
        if c != 1:
            cell.font = BLACK
        cell.border = BORD
        if r % 2 == 0:
            cell.fill = GREYF
    for c in (2, 3, 4):
        p.cell(row=r, column=c).number_format = NUM
    for c in (5, 6, 7, 8):
        p.cell(row=r, column=c).number_format = EUR
    p.cell(row=r, column=9).number_format = PCT

widths = [9, 13, 14, 13, 14, 14, 15, 24, 18]
for c, w in enumerate(widths, 1):
    p.column_dimensions[get_column_letter(c)].width = w
p.freeze_panes = "B4"

# Eckwerte
p["K3"] = "Eckwerte"
p["K3"].font = Font(name="Arial", size=11, bold=True)
marks = [("ARR nach 12 Monaten", "=F15"), ("ARR nach 24 Monaten", "=F27"),
         ("ARR nach 36 Monaten", "=F39"), ("Aktive Abos nach 36 Monaten", "=D39"),
         ("Deckungsbeitrag kumuliert (36 M.)", "=H39"),
         ("Marktdurchdringung nach 36 Monaten", "=I39")]
rr = 4
for lbl, fx in marks:
    p[f"K{rr}"] = lbl
    p[f"K{rr}"].font = BLACK
    p[f"L{rr}"] = fx
    p[f"L{rr}"].font = BOLD
    p[f"L{rr}"].number_format = PCT if "durchdringung" in lbl else (NUM if "Abos" in lbl else EUR)
    rr += 1
p.column_dimensions["K"].width = 34
p.column_dimensions["L"].width = 16

# ═══════════════════════════════════════════════════════════
# BLATT 3 — SZENARIEN
# ═══════════════════════════════════════════════════════════
s = wb.create_sheet("Szenarien")
s["A1"] = "Was bringt welche Kundenzahl?"
s["A1"].font = H1
s["A2"] = "Jahresumsatz bei unterschiedlichen Preispunkten und Abo-Zahlen."
s["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")

prices = [49, 79, 99, 149, 199]
counts = [50, 100, 250, 500, 1000, 1500]

s.cell(row=4, column=1, value="Abos ↓ / Preis →").font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
s.cell(row=4, column=1).fill = DARK
for i, pr in enumerate(prices):
    c = s.cell(row=4, column=2 + i, value=pr)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = DARK
    c.number_format = EUR2
    c.alignment = Alignment(horizontal="center")

for j, n in enumerate(counts):
    r = 5 + j
    c = s.cell(row=r, column=1, value=n)
    c.font = BOLD
    c.number_format = NUM
    c.fill = GREYF
    c.border = BORD
    for i in range(len(prices)):
        cell = s.cell(row=r, column=2 + i, value=f"=$A{r}*{get_column_letter(2+i)}$4*12")
        cell.number_format = EUR
        cell.font = BLACK
        cell.border = BORD

s.column_dimensions["A"].width = 20
for i in range(len(prices)):
    s.column_dimensions[get_column_letter(2 + i)].width = 15

s["A13"] = "Break-even-Betrachtung"
s["A13"].font = Font(name="Arial", size=11, bold=True)
s["A14"] = "Monatliche Fixkosten (Team, Technik, Recht)"
s["B14"] = 8000
s["B14"].font = BLUE
s["B14"].fill = YELLOW
s["B14"].number_format = EUR
s["A15"] = "Benötigte Abos für Kostendeckung"
s["B15"] = f"=ROUNDUP(B14/{P},0)"
s["B15"].font = BOLD
s["B15"].number_format = NUM
s["A16"] = "Entspricht Marktdurchdringung von"
s["B16"] = f"=B15/{TARGET}"
s["B16"].number_format = PCT
s["A17"] = "Auftragswert, ab dem sich das Abo für den Anbieter rechnet"
s["B17"] = f"={P}*12"
s["B17"].number_format = EUR
s["A18"] = "→ Ein einziger vermittelter Event-Auftrag (ca. 1.440 €) deckt das Jahresabo."
s["A18"].font = Font(name="Arial", size=10, italic=True, color="666666")
for rr2 in range(14, 18):
    s[f"A{rr2}"].font = BLACK

# ═══════════════════════════════════════════════════════════
# BLATT 4 — HINWEISE
# ═══════════════════════════════════════════════════════════
h = wb.create_sheet("Hinweise")
notes = [
    ("Quellen und Annahmen", 14, True),
    ("", 10, False),
    ("Belegte Zahlen", 11, True),
    ("5.330 Unternehmen im WZ 80.1 'Wach- und Sicherheitsdienste', Stichtag Sept. 2023.", 10, False),
    ("Quelle: Statistisches Bundesamt / Bewacherregister, Strukturerhebung Dienstleistungsbereich.", 9, False),
    ("Top 25 erwirtschaften 5,63 Mrd. € = rund 40 % des Marktvolumens ohne Geld- und Wertlogistik.", 10, False),
    ("Quelle: Lünendonk-Liste 2026, Führende Sicherheitsdienstleister in Deutschland.", 9, False),
    ("Branchenumsatz 2024: 14,02 Mrd. €. BDSW-Prognose 2025: rund 14,75 Mrd. €.", 10, False),
    ("", 10, False),
    ("Eigene Annahmen — nicht belegt, bitte selbst schärfen", 11, True),
    ("Adressierbarer Anteil 45 %: Betriebe mit Event- oder Kurzfristgeschäft. Langfristige", 10, False),
    ("Werkschutz-Rahmenverträge laufen über Ausschreibungen, nicht über einen Marktplatz.", 10, False),
    ("Churn 4 % pro Monat: typisch für KMU-SaaS mit monatlicher Kündbarkeit. Diese Zahl", 10, False),
    ("entscheidet über das gesamte Modell — bei 8 % halbiert sich der Lifetime Value.", 10, False),
    ("CAC 250 €: Mischkalkulation. Eigene Kaltakquise ist günstiger, Paid Ads teurer.", 10, False),
    ("", 10, False),
    ("Was das Modell nicht abbildet", 11, True),
    ("Die Nachfrageseite. Ohne Anfragen im System kündigen Anbieter nach zwei bis drei Monaten —", 10, False),
    ("dann ist die Churn-Annahme von 4 % deutlich zu optimistisch.", 10, False),
    ("Personalkosten, Entwicklung und Steuern. Der Deckungsbeitrag ist kein Gewinn.", 10, False),
    ("Jahresverträge. Bei 12-Monats-Bindung sinkt der Churn stark, die Abschlussquote aber auch.", 10, False),
]
for i, (t, sz, b) in enumerate(notes, start=1):
    c = h.cell(row=i, column=1, value=t)
    c.font = Font(name="Arial", size=sz, bold=b, color="0D1219" if b else "333333")
h.column_dimensions["A"].width = 100

wb.save(OUT)
print("gespeichert:", os.path.relpath(OUT, ROOT))
